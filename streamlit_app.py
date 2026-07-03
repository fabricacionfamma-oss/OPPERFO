import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# 1. CONFIGURACIÓN DE LA PÁGINA
# ==========================================
st.set_page_config(page_title="Panel Gerencial - Auditoría", layout="wide", page_icon="📊")

st.markdown("""
    <style>
    .main { background-color: #f8f9fa; }
    .stMetric { background-color: #ffffff; padding: 15px; border-radius: 10px; box-shadow: 0 2px 4px rgba(0,0,0,0.05); border-left: 5px solid #1A5276; }
    </style>
    """, unsafe_allow_html=True)

def color_performance(val):
    if val > 100 or val < 80:
        return 'color: #D32F2F; font-weight: bold;'
    else:
        return 'color: #2E7D32;'

# ==========================================
# 2. GENERADOR DE EXCEL (SIN GRÁFICOS)
# ==========================================
def generar_excel_descargable(df_resumen, df_dia, mes, anio):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    
    font_title = Font(name="Calibri", size=16, bold=True, color="1A5276")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Calibri", size=11)
    
    fill_header = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
    fill_alert_red = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    font_alert_red = Font(name="Calibri", size=11, bold=True, color="922B21")
    fill_alert_green = PatternFill(start_color="D4EFDF", end_color="D4EFDF", fill_type="solid")
    font_alert_green = Font(name="Calibri", size=11, color="196F3D")
    
    thin_border = Border(left=Side(style='thin', color='D5D8DC'), right=Side(style='thin', color='D5D8DC'),
                         top=Side(style='thin', color='D5D8DC'), bottom=Side(style='thin', color='D5D8DC'))
    align_center = Alignment(horizontal="center", vertical="center")
    
    # --- PESTAÑA 1: RESUMEN GENERAL ---
    ws1 = wb.active
    ws1.title = "Resumen General"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1.cell(row=2, column=2, value="REPORTE DE AUDITORÍA GERENCIAL").font = font_title
    ws1.cell(row=3, column=2, value=f"Periodo: Mes {mes} / Año {anio}").font = font_body
    
    headers_resumen = ["Operador", "Empresa", "Performance Mensual (%)"]
    for col_idx, header in enumerate(headers_resumen, start=2):
        c = ws1.cell(row=5, column=col_idx, value=header)
        c.font = font_header; c.fill = fill_header; c.alignment = align_center
        
    for row_idx, row_data in df_resumen.iterrows():
        r = row_idx + 6
        ws1.cell(row=r, column=2, value=row_data["Operador_Full"]).border = thin_border
        ws1.cell(row=r, column=3, value=row_data["Empresa"]).border = thin_border
        
        c_perf = ws1.cell(row=r, column=4, value=row_data["Perfo_Mensual (%)"]/100)
        c_perf.number_format = '0.0%'
        c_perf.border = thin_border
        
        val = row_data["Perfo_Mensual (%)"]
        if val < 80 or val > 100:
            c_perf.fill = fill_alert_red; c_perf.font = font_alert_red
        else:
            c_perf.fill = fill_alert_green; c_perf.font = font_alert_green

    for col in range(2, 5):
        ws1.column_dimensions[get_column_letter(col)].width = 30

    # --- PESTAÑA 2: DETALLE DIARIO (SOLO DATOS) ---
    ws2 = wb.create_sheet(title="Detalle Diario")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.cell(row=2, column=2, value="DESGLOSE DIARIO POR OPERADOR").font = font_title
    
    headers_diario = ["Operador", "Empresa", "Día", "Performance Diaria"]
    for col_idx, header in enumerate(headers_diario, start=2):
        c = ws2.cell(row=4, column=col_idx, value=header)
        c.font = font_header; c.fill = fill_header; c.alignment = align_center

    df_diario_sorted = df_dia.sort_values(["Operador_Full", "Dia"]).reset_index(drop=True)
    
    current_row = 5
    for _, row_data in df_diario_sorted.iterrows():
        ws2.cell(row=current_row, column=2, value=row_data["Operador_Full"]).border = thin_border
        ws2.cell(row=current_row, column=3, value=row_data["Empresa"]).border = thin_border
        ws2.cell(row=current_row, column=4, value=row_data["Dia"]).border = thin_border
        
        c_perf = ws2.cell(row=current_row, column=5, value=row_data["Perfo_SQL"]/100)
        c_perf.number_format = '0.0%'
        c_perf.border = thin_border
        
        val = row_data["Perfo_SQL"]
        if val < 80 or val > 100:
            c_perf.fill = fill_alert_red; c_perf.font = font_alert_red
        else:
            c_perf.fill = fill_alert_green; c_perf.font = font_alert_green
            
        current_row += 1

    for col in range(2, 6):
        ws2.column_dimensions[get_column_letter(col)].width = 25

    wb.save(output)
    output.seek(0)
    return output

# ==========================================
# 3. MOTOR SQL (AUTOMÁTICO)
# ==========================================
@st.cache_data(ttl=600)
def extraer_sql_data(mes, anio):
    def fetch_db(conn_name):
        try:
            conn = st.connection(conn_name, type="sql")
            q_m = f"""SELECT op.Name as Nombre, op.Docket as Legajo, 
                      SUM(p.Performance * p.ProductiveTime) / NULLIF(SUM(p.ProductiveTime), 0) as Perfo_SQL
                      FROM OPER_M_01 p JOIN OPERATOR op ON p.OperatorId = op.OperatorId 
                      WHERE p.Month = {mes} AND p.Year = {anio}
                      GROUP BY op.Name, op.Docket"""
            q_d = f"""SELECT op.Name as Nombre, op.Docket as Legajo, DAY(p.Date) as Dia, 
                      SUM(p.Performance * p.ProductiveTime) / NULLIF(SUM(p.ProductiveTime), 0) as Perfo_SQL
                      FROM OPER_D_01 p JOIN OPERATOR op ON p.OperatorId = op.OperatorId 
                      WHERE MONTH(p.Date) = {mes} AND YEAR(p.Date) = {anio}
                      GROUP BY op.Name, op.Docket, DAY(p.Date)"""
            return conn.query(q_m), conn.query(q_d)
        except Exception:
            return pd.DataFrame(), pd.DataFrame()

    df_m_fa, df_d_fa = fetch_db("famma_db")
    df_m_fu, df_d_fu = fetch_db("fumi_db")
    
    # Asignar empresa también a los DataFrames diarios
    if not df_m_fa.empty: df_m_fa['Empresa'] = 'FAMMA'
    if not df_d_fa.empty: df_d_fa['Empresa'] = 'FAMMA'
    if not df_m_fu.empty: df_m_fu['Empresa'] = 'FUMISCOR'
    if not df_d_fu.empty: df_d_fu['Empresa'] = 'FUMISCOR'
    
    df_mes = pd.concat([df_m_fa, df_m_fu], ignore_index=True)
    df_dia = pd.concat([df_d_fa, df_d_fu], ignore_index=True)
    
    if not df_mes.empty:
        df_mes = df_mes[~df_mes['Legajo'].astype(str).str.upper().str.startswith('FW')]
    if not df_dia.empty:
        df_dia = df_dia[~df_dia['Legajo'].astype(str).str.upper().str.startswith('FW')]
    
    for df in [df_mes, df_dia]:
        if not df.empty:
            df['Perfo_SQL'] = np.where(df['Perfo_SQL'] > 1.5, df['Perfo_SQL']/100, df['Perfo_SQL']) * 100
            df['Operador_Full'] = df['Nombre'].astype(str).str.upper() + " (" + df['Legajo'].astype(str) + ")"
            
    return df_mes.drop_duplicates(subset=['Operador_Full']), df_dia

# ==========================================
# 4. BARRA LATERAL
# ==========================================
st.sidebar.header("📅 Periodo a Auditar")
mes_sel = st.sidebar.slider("Mes", 1, 12, 4)
anio_sel = st.sidebar.number_input("Año", 2024, 2030, 2026)

# ==========================================
# 5. PROCESAMIENTO Y DASHBOARD
# ==========================================
with st.spinner("Sincronizando con base de datos SQL..."):
    df_sql_mes, df_sql_dia = extraer_sql_data(mes_sel, anio_sel)

st.title("🏭 Auditoría Gerencial Wiidem")

if not df_sql_mes.empty:
    
    st.header("🏆 Resumen General de Operarios")
    
    df_resumen = df_sql_mes[['Operador_Full', 'Empresa', 'Perfo_SQL']].copy()
    df_resumen = df_resumen.rename(columns={'Perfo_SQL': 'Perfo_Mensual (%)'})
    df_resumen = df_resumen.sort_values('Perfo_Mensual (%)', ascending=False).reset_index(drop=True)

    col_export, col_vacio = st.columns([1, 4])
    with col_export:
        archivo_excel = generar_excel_descargable(df_resumen, df_sql_dia, mes_sel, anio_sel)
        st.download_button(
            label="📥 Descargar Reporte Excel",
            data=archivo_excel,
            file_name=f"Auditoria_Gerencial_{mes_sel}_{anio_sel}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
    
    with st.expander("Ver Listado Completo con Alertas (Rojo <80% o >100%)", expanded=True):
        evento = st.dataframe(
            df_resumen.style.map(color_performance, subset=['Perfo_Mensual (%)'])
            .format({'Perfo_Mensual (%)': '{:.1f}%'}),
            use_container_width=True, hide_index=True,
            on_select="rerun",
            selection_mode="single-row"
        )

        if len(evento.selection.rows) > 0:
            fila_seleccionada = evento.selection.rows[0]
            op_click = df_resumen.iloc[fila_seleccionada]['Operador_Full']
            st.session_state['operador_seleccionado'] = op_click

    st.divider()
    
    st.header("🔍 Auditoría Individual")
    
    lista_operadores = sorted(df_resumen['Operador_Full'].unique())

    if 'operador_seleccionado' in st.session_state and st.session_state['operador_seleccionado'] in lista_operadores:
        indice_default = lista_operadores.index(st.session_state['operador_seleccionado'])
    else:
        indice_default = 0

    op_sel = st.selectbox(
        "Seleccione Operador para analizar detalle diario:", 
        lista_operadores,
        index=indice_default
    )
    
    st.session_state['operador_seleccionado'] = op_sel
    
    sql_op = df_sql_dia[df_sql_dia['Operador_Full'] == op_sel].copy()
    df_dash = sql_op[['Dia', 'Perfo_SQL']].copy().fillna(0)
    df_dash['Fecha_Label'] = df_dash['Dia'].astype(str) + f"/{mes_sel}"

    c1, c2 = st.columns(2)
    with c1:
        excluidos = st.multiselect("Eliminar días atípicos del promedio:", df_dash['Fecha_Label'].tolist())
    with c2:
        multiplicador = st.number_input("Multiplicador de Ajuste (%)", 10.0, 200.0, 100.0, 5.0)

    df_v = df_dash[~df_dash['Fecha_Label'].isin(excluidos)]
    p_orig = df_dash['Perfo_SQL'].mean() if not df_dash.empty else 0
    p_base = df_v['Perfo_SQL'].mean() if not df_v.empty else 0
    p_final = p_base * (multiplicador / 100.0)

    m1, m2, m3 = st.columns(3)
    m1.metric("Perfo Mensual Original", f"{p_orig:.1f}%")
    m2.metric(f"Perfo Días Válidos", f"{p_base:.1f}%", f"{p_base-p_orig:+.1f}%")
    m3.metric("🎯 PERFO FINAL AUDITADA", f"{p_final:.1f}%", f"x {multiplicador/100:.2f}")

    if not df_dash.empty:
        df_dash['Estado'] = df_dash['Fecha_Label'].apply(lambda x: 'Excluido' if x in excluidos else 'Válido')
        st.plotly_chart(px.bar(df_dash, x='Fecha_Label', y='Perfo_SQL', color='Estado', 
                               color_discrete_map={'Válido':'#1A5276', 'Excluido':'#D5D8DC'},
                               text=df_dash['Perfo_SQL'].apply(lambda x: f"{x:.1f}%")), use_container_width=True)
    else:
        st.warning("No hay datos diarios para este operador en el mes seleccionado.")

else:
    st.info("No se encontraron datos en la base de datos para el mes y año seleccionados.")
